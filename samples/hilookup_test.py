"""
The origin of the test datasource can be found on
https://www.foodstandards.gov.au/science/monitoringnutrients/afcd/Pages/downloadableexcelfiles.aspx

Required packages:
- pandas, openpyxl, fuzzywuzzy, python-Levenshtein
"""


import pandas as pd
import sys
import os
import time
import threading


if __name__=="__main__":
    # check if the second argv passed by the command line
    # if  yes, expect the index of trg row to replace the rownum_trg_todebug_list
    trg_rownum_todebug_list = [0]
    src_rownum_todebug_list = [9]
    if len(sys.argv[1:]) > 0:
        for row in sys.argv[1].split(','):
            trg_rownum_todebug_list.append(int(row))
        if len(sys.argv[1:]) > 1:
            for row in sys.argv[2].split(','):
                src_rownum_todebug_list.append(int(row))

    # define dump directory
    #dump_directory = os.environ['TEMP']
    dump_directory =  r"E:\hilookup_samples"
    
    # define src data details
    # -----------------------
    src_path = r"E:\hilookup_samples\Release 1 - Food details file.xlsx"
    src_tablename = "Release 1 - Food File"
    src_df = pd.read_excel(src_path, src_tablename, engine='openpyxl')



    # define trg data details
    # ------------------------
    trg_path = r"E:\hilookup_samples\target_data.xlsx"
    trg_tablename = "Sheet1"
    trg_df = pd.read_excel(trg_path, trg_tablename, engine='openpyxl')
    trg_df['_rownum'] = trg_df.index


    # how many rows availables
    print("num of trg_df records: {}".format(len(trg_df)))
    print("num of src_df records: {}".format(len(src_df)))


    # define numof_output
    numof_output = 5

    import hilookup
    hi = hilookup.HILookup(src_df,trg_df,numof_output)

    # set parameters
    hi.fuzzratio_min = 60
    hi.src_fieldname_toevaluate_list = ['Name']
    hi.src_wordindex_group_dict = {",":"left-to-right"}
    hi.src_wordindex_simple = "left-to-right"
    hi.trg_fieldname_toevaluate_list = ['Name']
    hi.trg_wordindex_simple = "left-to-right"

    #--------------------------
    # define sync or unsync
    #--------------------------
    from hilookup.helper import ProgressBar
    _async = True
    if _async == True:
        thrd = threading.Thread(target=hi.hilookup,args=())
        thrd.start()

        # create a progress bar so we can show it when we loop into the live thread

        item_total = len(hi.trg_df)
        pb = ProgressBar(item_total)

        while thrd.is_alive():
            # show the progress bar
            pb.done_tasks = hi.trg_processed_cnt
            print(pb.get_progress(), "scanned", hi.scans, "matches",len(hi.trg_matching_list), end="\r")
            # take a sleep for few seconds
            time.sleep(5)

        thrd.join()
        # show the final state of the progress bar
        pb.done_tasks = hi.trg_processed_cnt
        print(pb.get_progress(), "scanned", hi.scans, "matches",len(hi.trg_matching_list))

    else:
        hi.hilookup()

    # ---------------
    # process results
    # ---------------

    hi.trg_df["_rank"] = -1 # set -1 for each trg row so we can sort it Smallest to Largest

    for row_um in hi.trg_matching_list:
        for i in range(0,numof_output):
            if i < len(row_um.matched_src_list):
                group_item_dict = {}
                _idx_src = row_um.matched_src_list[i].rowid
                _score_weighted = row_um.matched_src_list[i].score_weighted
                _penalty = row_um.matched_src_list[i].penalty
                _net_score = int(_score_weighted - _penalty)

                debug = False
                if debug:
                    print("-------------debug--------------")
                    print('trg_rowid {}'.format(row_um.rowid))
                    print('src_rowid {}'.format(row_um.matched_src_list[i].rowid))
                    print("score {} = score_weighted {} penalty {}".format(_score_weighted - _penalty,_score_weighted,_penalty))

                # create a dict to hold result then add to trg_df
                src_dict = {}
                src_dict["_rownum"] = row_um.rowid # the trgrowid
                src_dict["_src_rowid"] = row_um.matched_src_list[i].rowid
                src_dict["score"] = row_um.matched_src_list[i].score_weighted - row_um.matched_src_list[i].penalty
                src_dict["_rank"] = i
                # loop the src column and populate the src_dict                
                src_cols = hi.src_df.columns.tolist()
                for col in src_cols:
                    src_dict[col] = hi.src_df.at[row_um.matched_src_list[i].rowid, col]
                    
                hi.trg_df = hi.trg_df.append(src_dict, ignore_index=True)    
                


    # rename _rownum column as group
    hi.trg_df = hi.trg_df.rename(columns={"_rownum":"group","_rowid":"_trg_rowid"})
    # reorder the column name
    cols = hi.trg_df.columns.tolist()
    cols_neworder = []
    cols_neworder.append('group')
    for col in cols:
        if col != 'group':
            cols_neworder.append(col)
    hi.trg_df = hi.trg_df[cols_neworder]


    # use openpyxl to enable excel formatting
    # we want to format blue to any background cell belong to the source
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font
    # create a openpyxl workbook and add a worksheet in it and set it as the active ws
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1' # name it as a common name
    # copy all the rows from pandas to to the ws
    for r in dataframe_to_rows(hi.trg_df, index=False, header=True):
        ws.append(r)
    # optional - format the header as the Pandas style
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    # iterate the ws
    min_row = len(trg_df) + 2 # 2 is made from 1 for header and another 1 because pandas starts with 0
    max_row = len(hi.trg_df) + 1 # 1 is the header.
    for rows in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1):
        for cell in rows:
            # color blue all of the row    
            cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type = "solid")    
    
    # save the wb into an excel file.
    result_path = os.path.join(dump_directory,'results_' + str(int(time.time())) + '.xlsx')
    wb.save(result_path) 